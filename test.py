import torch
import torch.nn as nn
from torch.utils.data import DataLoader, Dataset
import numpy as np
from dotmap import DotMap
import wandb
from wandb.wandb_run import Run
import openpyxl
from openpyxl.styles import Alignment
import pickle
from PIL import Image
from pathlib import Path
from typing import List, Tuple
from datetime import datetime
from einops import rearrange
from sklearn.linear_model import Ridge
from scipy import stats
import argparse

from data import LIVEDataset, CSIQDataset, TID2013Dataset, KADID10KDataset, FLIVEDataset, SPAQDataset
from utils.utils import PROJECT_ROOT, parse_command_line_args, merge_configs, parse_config
from models.simclr import SimCLR


synthetic_datasets = ["live", "csiq", "tid2013", "kadid10k"]
authentic_datasets = ["flive", "spaq"]


def test(args: DotMap,
         model: nn.Module,
         logger: Run,
         device: torch.device) -> None:
    """
    Test pretrained model on the test datasets. Performs a grid search over the validation splits to find the best
    alpha value for the regression for each dataset. Saves a CSV file with the results and a pickle file with the
    regressor for each dataset.

    Args:
        args (dotmap.DotMap): test arguments
        model (torch.nn.Module): model to test
        logger (wandb.wandb_run.Run): wandb logger
        device (torch.device): device to use for testing
    """
    checkpoint_base_path = PROJECT_ROOT / "experiments"
    checkpoint_path = checkpoint_base_path / args.experiment_name
    regressor_path = checkpoint_path / "regressors"
    regressor_path.mkdir(parents=True, exist_ok=True)

    eval_type = args.get("eval_type", "scratch")

    model.eval()

    srocc_all, plcc_all, regressors, alphas, best_worst_results_all = get_results(model=model,
                                                                                  data_base_path=args.data_base_path,
                                                                                  datasets=args.test.datasets,
                                                                                  num_splits=args.test.num_splits,
                                                                                  phase="test",
                                                                                  alpha=args.test.alpha,
                                                                                  grid_search=args.test.grid_search,
                                                                                  crop_size=args.test.crop_size,
                                                                                  batch_size=args.test.batch_size,
                                                                                  num_workers=args.test.num_workers,
                                                                                  device=device, eval_type=eval_type)

    # Compute the median for each list in srocc_all and plcc_all
    srocc_all_median = {key: np.median(value["global"]) for key, value in srocc_all.items()}
    plcc_all_median = {key: np.median(value["global"]) for key, value in plcc_all.items()}

    # Compute the synthetic and authentic averages
    srocc_synthetic_avg = np.mean(
        [srocc_all_median[key] for key in srocc_all_median.keys() if key in synthetic_datasets])
    plcc_synthetic_avg = np.mean([plcc_all_median[key] for key in plcc_all_median.keys() if key in synthetic_datasets])
    srocc_authentic_avg = np.mean(
        [srocc_all_median[key] for key in srocc_all_median.keys() if key in authentic_datasets])
    plcc_authentic_avg = np.mean([plcc_all_median[key] for key in plcc_all_median.keys() if key in authentic_datasets])

    # Compute the global average
    srocc_avg = np.mean(list(srocc_all_median.values()))
    plcc_avg = np.mean(list(plcc_all_median.values()))

    print(f"{'Dataset':<15} {'Alpha':<15} {'SROCC':<15} {'PLCC':<15}")
    for dataset in srocc_all_median.keys():
        print(f"{dataset:<15} {alphas[dataset]} {srocc_all_median[dataset]:<15.4f} {plcc_all_median[dataset]:<15.4f}")
    print(f"{'Synthetic avg':<15} {srocc_synthetic_avg:<15.4f} {plcc_synthetic_avg:<15.4f}")
    print(f"{'Authentic avg':<15} {srocc_authentic_avg:<15.4f} {plcc_authentic_avg:<15.4f}")
    print(f"{'Global avg':<15} {srocc_avg:<15.4f} {plcc_avg:<15.4f}")

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Create a sheet for the median SROCC values
    median_sheet = workbook.create_sheet('Median', 0)
    median_sheet.append(['Dataset', 'Alpha', 'SROCC', 'PLCC'])
    for dataset, srocc_results in srocc_all.items():
        median_sheet.append([dataset, alphas[dataset], srocc_all_median[dataset], plcc_all_median[dataset]])
    median_sheet.append(['Synthetic avg', '', srocc_synthetic_avg, plcc_synthetic_avg])
    median_sheet.append(['Authentic avg', '', srocc_authentic_avg, plcc_authentic_avg])
    median_sheet.append(['Global avg', '',  srocc_avg, plcc_avg])

    # Create a sheet for each dataset with all the SROCC and PLCC values
    for (dataset, srocc_results), (_, plcc_results) in zip(srocc_all.items(), plcc_all.items()):
        sheet = workbook.create_sheet(dataset)
        # Write header
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        sheet.cell(row=1, column=1).value = "Split"
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        for i in range(1, len(srocc_results) + 1):
            sheet.merge_cells(start_row=1, start_column=(i * 2), end_row=1, end_column=(i * 2) + 1)
            sheet.cell(row=1, column=(i * 2)).value = list(srocc_results.keys())[i - 1]
            sheet.cell(row=1, column=(i * 2)).alignment = Alignment(horizontal='center')
            sheet.cell(row=2, column=(i * 2)).value = "SROCC"
            sheet.cell(row=2, column=(i * 2) + 1).value = "PLCC"

        for i in range(len(srocc_results["global"])):
            row = [i]
            for dist_type in srocc_results.keys():
                row += [srocc_results[dist_type][i]]
                row += [plcc_results[dist_type][i]]
            sheet.append(row)

    # Set the median sheet as the active sheet
    workbook.active = 0

    # Remove default sheet
    workbook.remove(workbook['Sheet'])

    # Save the workbook to a file
    workbook.save(checkpoint_path / 'results.xlsx')

    if logger:
        for dataset in srocc_all_median.keys():
            logger.summary[f"test_srocc_{dataset}"] = srocc_all_median[dataset]
            logger.summary[f"test_plcc_{dataset}"] = plcc_all_median[dataset]
            logger.summary[f"test_alpha_{dataset}"] = alphas[dataset]
        logger.summary[f"test_srocc_synthetic_avg"] = srocc_synthetic_avg
        logger.summary[f"test_plcc_synthetic_avg"] = plcc_synthetic_avg
        logger.summary[f"test_srocc_authentic_avg"] = srocc_authentic_avg
        logger.summary[f"test_plcc_authentic_avg"] = plcc_authentic_avg
        logger.summary[f"test_srocc_avg"] = srocc_avg
        logger.summary[f"test_plcc_avg"] = plcc_avg

        # Log best and worst results
        for dataset in best_worst_results_all.keys():
            for key in best_worst_results_all[dataset].keys():  # key is either "best" or "worst"
                column_names = ["Image", "Predicted", "GT", "Difference", "Path"]
                table_data = []
                for i in range(len(best_worst_results_all[dataset][key]["images"])):
                    img_path = best_worst_results_all[dataset][key]["images"][i]
                    img = wandb.Image(Image.open(img_path))
                    gt = best_worst_results_all[dataset][key]["gts"][i]
                    pred = best_worst_results_all[dataset][key]["preds"][i]
                    diff = pred - gt
                    table_data.append([img, pred, gt, diff, str(img_path)])
                logger.log({f"test_{dataset}_{key}_results": wandb.Table(data=table_data, columns=column_names)})

    for dataset, regressor in regressors.items():
        filename = f"{dataset}_srocc_{srocc_all_median[dataset]:.4f}_plcc_{plcc_all_median[dataset]:.4f}.pkl"
        with open(regressor_path / filename, "wb") as f:
            pickle.dump(regressor, f)


def get_results(model: nn.Module,
                data_base_path: Path,
                datasets: List[str],
                num_splits: int,
                phase: str,
                alpha: float,
                grid_search: bool,
                crop_size: int,
                batch_size: int,
                num_workers: int,
                device: torch.device,
                eval_type: str = "scratch") -> Tuple[dict, dict, dict, dict, dict]:
    """
    Get the results for the given model and datasets. Depending on the phase parameter, can be used both for validation
    and test. If phase == 'test' and grid_search == True, performs a grid search over the validation splits to find the best
    alpha value for the regression for each dataset. The results related to synthetic datasets contain also the results
    for each distortion type.

    Args:
        model (torch.nn.Module): model to test
        data_base_path (pathlib.Path): base path of the datasets
        datasets (list): list of datasets
        num_splits (int): number of splits
        phase (str): phase of the datasets. Must be in ['val', 'test']
        alpha (float): alpha value to use for regression. During test, if None, performs a grid search
        grid_search (bool): whether to perform a grid search over the validation splits to find the best alpha value for the regression
        crop_size (int): crop size
        batch_size (int): batch size
        num_workers (int): number of workers for the dataloaders
        device (torch.device): device to use for testing
        eval_type (str): Whether to test a model trained from scratch or the one pretrained by the authors of the ARNIQA paper.

    Returns:
        srocc_all (dict): dictionary containing the SROCC results
        plcc_all (dict): dictionary containing the PLCC results
        regressors (dict): dictionary containing the regressors
        alphas (dict): dictionary containing the alpha values used for the regression
        best_worst_results_all (dict): dictionary containing the best and worst results
    """
    srocc_all = {}
    plcc_all = {}
    regressors = {}
    alphas = {}
    best_worst_results_all = {}

    assert phase in ["val", "test"], "Phase must be in ['val', 'test']"

    print(f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')} - Starting {phase} phase")
    for d in datasets:
        if d == "live":
            dataset = LIVEDataset(data_base_path / "LIVE", phase="all", crop_size=crop_size)
            dataset_num_splits = num_splits
            dataset_name = "LIVE"
        elif d == "csiq":
            dataset = CSIQDataset(data_base_path / "CSIQ", phase="all", crop_size=crop_size)
            dataset_num_splits = num_splits
            dataset_name = "CSIQ"
        elif d == "tid2013":
            dataset = TID2013Dataset(data_base_path / "TID2013", phase="all", crop_size=crop_size)
            dataset_num_splits = num_splits
            dataset_name = "TID2013"
        elif d == "kadid10k":
            dataset = KADID10KDataset(data_base_path / "KADID10K", phase="all", crop_size=crop_size)
            dataset_num_splits = num_splits
            dataset_name = "KADID-10K"
        elif d == "flive":
            dataset = FLIVEDataset(data_base_path / "FLIVE", phase="all", crop_size=crop_size)
            dataset_num_splits = 1
            dataset_name = "FLIVE"
        elif d == "spaq":
            dataset = SPAQDataset(data_base_path / "SPAQ", phase="all", crop_size=crop_size)
            dataset_num_splits = num_splits
            dataset_name = "SPAQ"
        else:
            raise ValueError(f"Dataset {d} not supported")

        srocc_dataset, plcc_dataset, regressor, alpha, best_worst_results = compute_metrics(model, dataset,
                                                                                            dataset_num_splits, phase,
                                                                                            alpha, grid_search,
                                                                                            batch_size, num_workers,
                                                                                            device, eval_type)
        srocc_all[d] = srocc_dataset
        plcc_all[d] = plcc_dataset
        regressors[d] = regressor
        alphas[d] = alpha
        best_worst_results_all[d] = best_worst_results
        print(f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')} - {dataset_name}:"
              f" SRCC: {np.median(srocc_dataset['global']):.3f} - PLCC: {np.median(plcc_dataset['global']):.3f}")

    return srocc_all, plcc_all, regressors, alphas, best_worst_results_all


def compute_metrics(model: nn.Module,
                    dataset: Dataset,
                    num_splits: int,
                    phase: str,
                    alpha: float,
                    grid_search: bool,
                    batch_size: int,
                    num_workers: int,
                    device: torch.device,
                    eval_type: str = "scratch") -> Tuple[dict, dict, Ridge, float, dict]:
    """
    Compute the metrics for the given model and dataset. If phase == 'test' and grid_search == True, performs a grid search
    over the validation splits to find the best alpha value for the regression.

    Args:
        model (torch.nn.Module): model to test
        dataset (torch.utils.data.Dataset): dataset to test on
        num_splits (int): number of splits
        phase (str): phase of the datasets. Must be in ['val', 'test']
        alpha (float): alpha value to use for regression. During test, if None, performs a grid search
        grid_search (bool): whether to perform a grid search over the validation splits to find the best alpha value for the regression
        batch_size (int): batch size
        num_workers (int): number of workers for the dataloaders
        device (torch.device): device to use for testing
        eval_type (str): Whether to test a model trained from scratch or the one pretrained by the authors of the ARNIQA paper.

    Returns:
        srocc_dataset (dict): dictionary containing the SROCC results for the dataset
        plcc_dataset (dict): dictionary containing the PLCC results for the dataset
        regressor (Ridge): Ridge regressor
        alpha (float): alpha value used for the regression
        best_worst_results (dict): dictionary containing the best and worst results
    """
    srocc_dataset = {"global": []}
    plcc_dataset = {"global": []}
    best_worst_results = {}     # Best and worst 16 results according to the difference between the predicted and the true MOS
    dist_types = None
    if dataset.is_synthetic:
        dist_types = set(dataset.distortion_types)
        for dist_type in dist_types:
            srocc_dataset[dist_type] = []
            plcc_dataset[dist_type] = []

    dataloader = DataLoader(dataset, batch_size=batch_size, shuffle=False, num_workers=num_workers, pin_memory=True)

    features, scores = get_features_scores(model, dataloader, device, eval_type)

    # Perform grid search over the validation splits to find the best alpha value for the regression
    if phase == "test" and grid_search:
        best_alpha = alpha_grid_search(dataset=dataset, features=features, scores=scores, num_splits=num_splits)
    else:
        best_alpha = alpha

    for i in range(num_splits):
        train_indices = dataset.get_split_indices(split=i, phase="train")
        test_indices = dataset.get_split_indices(split=i, phase=phase)

        dist_indices = None
        if dataset.is_synthetic:
            dist_indices = {dist_type: np.where(dataset.distortion_types[test_indices] == dist_type)[0] for dist_type in dist_types}

        # for each index generate 5 indices (one for each crop)
        train_indices = np.repeat(train_indices * 5, 5) + np.tile(np.arange(5), len(train_indices))
        test_indices = np.repeat(test_indices * 5, 5) + np.tile(np.arange(5), len(test_indices))

        train_features = features[train_indices]
        train_scores = scores[train_indices]

        regressor = Ridge(alpha=best_alpha).fit(train_features, train_scores)

        test_features = features[test_indices]
        test_scores = scores[test_indices]
        test_scores = test_scores[::5]  # Scores are repeated for each crop, so we only keep the first one
        orig_test_indices = test_indices[::5] // 5  # Get original indices

        preds = regressor.predict(test_features)
        preds = np.mean(np.reshape(preds, (-1, 5)), 1)  # Average the predictions of the 5 crops of the same image

        srocc_dataset["global"].append(stats.spearmanr(preds, test_scores)[0])
        plcc_dataset["global"].append(stats.pearsonr(preds, test_scores)[0])

        if dataset.is_synthetic:
            for dist_type in dist_types:
                srocc_dataset[dist_type].append(stats.spearmanr(preds[dist_indices[dist_type]], test_scores[dist_indices[dist_type]])[0])
                plcc_dataset[dist_type].append(stats.pearsonr(preds[dist_indices[dist_type]], test_scores[dist_indices[dist_type]])[0])

        # Compute best and worst results
        if i == 0:
            diff = np.abs(preds - test_scores)
            sorted_diff_indices = np.argsort(diff)
            best_indices = sorted_diff_indices[:16]
            worst_indices = sorted_diff_indices[-16:][::-1]
            best_worst_results["best"] = {"images": dataset.images[orig_test_indices[best_indices]], "gts": test_scores[best_indices], "preds": preds[best_indices]}
            best_worst_results["worst"] = {"images": dataset.images[orig_test_indices[worst_indices]], "gts": test_scores[worst_indices], "preds": preds[worst_indices]}

    # Train a regressor on the whole dataset for saving purposes
    regressor = Ridge(alpha=best_alpha).fit(features, scores)

    return srocc_dataset, plcc_dataset, regressor, best_alpha, best_worst_results


def get_features_scores(model: nn.Module,
                        dataloader: DataLoader,
                        device: torch.device,
                        eval_type: str = "scratch") -> Tuple[np.ndarray, np.ndarray]:
    """
    Get the features and scores for the given model and dataloader.

    Args:
        model (torch.nn.Module): model to test
        dataloader (torch.utils.data.Dataloader): dataloader
        device (torch.device): device to use for testing
        eval_type (str): Whether to test a model trained from scratch or the one pretrained by the authors of the ARNIQA paper.

    Returns:
        features (np.ndarray): features
        scores (np.ndarray): ground-truth MOS scores
    """
    feats = np.zeros((0, model.encoder.feat_dim * 2))   # Double the features because of the original and downsampled image
    scores = np.zeros(0)

    for i, batch in enumerate(dataloader):
        img_orig = batch["img"].to(device)
        img_ds = batch["img_ds"].to(device)
        mos = batch["mos"]

        img_orig = rearrange(img_orig, "b n c h w -> (b n) c h w")
        img_ds = rearrange(img_ds, "b n c h w -> (b n) c h w")
        mos = mos.repeat_interleave(5)  # repeat MOS for each crop

        with torch.cuda.amp.autocast(), torch.no_grad():
            if eval_type == "scratch":
                f_orig, _ = model(img_orig)
                f_ds, _ = model(img_ds)
                f = torch.hstack((f_orig, f_ds))
            elif eval_type == "arniqa":
                _, f = model(img_orig, img_ds, return_embedding=True)

        feats = np.concatenate((feats, f.cpu().numpy()), 0)
        scores = np.concatenate((scores, mos.numpy()), 0)

    return feats, scores


def alpha_grid_search(dataset: Dataset,
                      features: np.ndarray,
                      scores: np.ndarray,
                      num_splits: int) -> float:
    """
    Perform a grid search over the validation splits to find the best alpha value for the regression based on the SROCC
    metric. The grid search is performed over the range [1-e3, 1e3, 100].

    Args:
        dataset (Dataset): dataset to use
        features (np.ndarray): features extracted with the model to test
        scores (np.ndarray): ground-truth MOS scores
        num_splits (int): number of splits to use

    Returns:
        alpha (float): best alpha value
    """

    grid_search_range = [1e-3, 1e3, 100]
    alphas = np.geomspace(*grid_search_range, endpoint=True)
    srocc_all = [[] for _ in range(len(alphas))]

    for i in range(num_splits):
        train_indices = dataset.get_split_indices(split=i, phase="train")
        val_indices = dataset.get_split_indices(split=i, phase="val")

        # for each index generate 5 indices (one for each crop)
        train_indices = np.repeat(train_indices * 5, 5) + np.tile(np.arange(5), len(train_indices))
        val_indices = np.repeat(val_indices * 5, 5) + np.tile(np.arange(5), len(val_indices))

        train_features = features[train_indices]
        train_scores = scores[train_indices]

        val_features = features[val_indices]
        val_scores = scores[val_indices]
        val_scores = val_scores[::5]  # Scores are repeated for each crop, so we only keep the first one

        for idx, alpha in enumerate(alphas):
            regressor = Ridge(alpha=alpha).fit(train_features, train_scores)
            preds = regressor.predict(val_features)
            preds = np.mean(np.reshape(preds, (-1, 5)), 1)  # Average the predictions of the 5 crops of the same image
            srocc_all[idx].append(stats.spearmanr(preds, val_scores)[0])

    srocc_all_median = [np.median(srocc) for srocc in srocc_all]
    srocc_all_median = np.array(srocc_all_median)
    best_alpha_idx = np.argmax(srocc_all_median)
    best_alpha = alphas[best_alpha_idx]

    return best_alpha


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--config', type=str, required=True, help='Configuration file')
    parser.add_argument("--eval_type", type=str, default="scratch", choices=["scratch", "arniqa"],
                        help="Whether to test a model trained from scratch or the one pretrained by the authors of the"
                             "paper. Must be in ['scratch', 'arniqa']")
    args, _ = parser.parse_known_args()
    eval_type = args.eval_type
    config = parse_config(args.config)
    args = parse_command_line_args(config)
    args = merge_configs(config, args)
    args.eval_type = eval_type
    args.data_base_path = Path(args.data_base_path)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    if args.eval_type == "scratch":
        model = SimCLR(encoder_params=args.model.encoder, temperature=args.model.temperature)
        checkpoint_base_path = PROJECT_ROOT / "experiments"
        assert (checkpoint_base_path / args.experiment_name).exists(), \
            f"Experiment {(checkpoint_base_path / args.experiment_name)} does not exist"
        checkpoint_path = checkpoint_base_path / args.experiment_name / "pretrain"
        checkpoint_path = [ckpt_path for ckpt_path in checkpoint_path.glob("*.pth") if "best" in ckpt_path.name][0]
        checkpoint = torch.load(checkpoint_path, map_location="cpu")
        model.load_state_dict(checkpoint, strict=True)
    elif args.eval_type == "arniqa":
        model = torch.hub.load(repo_or_dir="miccunifi/ARNIQA", source="github", model="ARNIQA")
    else:
        raise ValueError(f"Eval type {args.eval_type} not supported")
    model.to(device)
    model.eval()

    if args.logging.use_wandb:
        logger = wandb.init(project=args.logging.wandb.project,
                            entity=args.logging.wandb.entity,
                            name=args.experiment_name,
                            config=args,
                            mode="online" if args.logging.wandb.online else "offline")
    else:
        logger = None

    test(args, model, logger, device)

